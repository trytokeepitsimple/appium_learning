import pytest
import openpyxl

def get_data():
    # return[
    #     #mainList
    #
    #
    #     ["Delhi" , "India"],
    #     ["Dubai", "UAE"]
    #     #dataList
    # ]
    workbook= openpyxl.load_workbook("/home/devesh-ssd/PycharmProjects/appium_testing/excel/Financial Sample.xlsx")
    sheet = workbook["login_details"]
    rows = sheet.max_row
    cols = sheet.max_column

    #pulling data from sheet
    mainList = []
    for row in range(2,rows+1):
        dataList = []
        for col in range(1,cols+1):
            data = sheet.cell(row=row,column=col).value
            dataList.insert(col,data)
        mainList.insert(row,dataList)
    print(mainList)
    return mainList



data = get_data()
print(data)

# @pytest.mark.parametrize("username,password",get_data())
# def test_do_login(username,password):
#     print(username,password)