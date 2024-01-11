import openpyxl

workbook = openpyxl.load_workbook("/home/devesh-ssd/PycharmProjects/appium_testing/excel/Financial Sample.xlsx")


#name of sheet from where one need to read data
sheet = workbook["login_details"]


#max rows and max cols
total_rows = sheet.max_row
total_cols = sheet.max_column

# print("Total columns : ",str(total_cols))
# print("Total Rows : ",str(total_rows))


#print data from specific cells
# print(sheet.cell(row=2,column=1).value)


#read data from whole sheet

# for rows in range(1, total_rows+1):
#     for cols in range(1, total_cols+1):
#         print(sheet.cell(row=rows, column=cols).value, end="  ")
#     #empty print to create entries in new line or row
#     print()



#write data into sheet


#set value in a specific cell
sheet.cell(row=3,column=1).value="ghi"
sheet.cell(row=3,column=2).value="pass3"
#after everything save the excel file
workbook.save("/home/devesh-ssd/PycharmProjects/appium_testing/excel/Financial Sample.xlsx")