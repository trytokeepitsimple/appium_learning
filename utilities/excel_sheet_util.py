import openpyxl

def get_rows_count(path, sheet):
    workbook = openpyxl.load_workbook(path)

    # name of sheet from where one need to read data
    sheet = workbook[sheet]
    return sheet.max_row


def get_cols_count(path,sheet):
    workbook = openpyxl.load_workbook(path)

    # name of sheet from where one need to read data
    sheet = workbook[sheet]
    return sheet.max_column


def get_data(path, sheet , rowNum , colNum):
    workbook = openpyxl.load_workbook(path)

    # name of sheet from where one need to read data
    sheet = workbook[sheet]
    return sheet.cell(row=rowNum,column=colNum).value


def set_data(path, sheet , rowNum , colNum , data):
    workbook = openpyxl.load_workbook(path)

    # name of sheet from where one need to read data
    sheet = workbook[sheet]
    sheet.cell(row=rowNum, column=colNum).value=data
    workbook.save(path)

path='/home/devesh-ssd/PycharmProjects/appium_testing/excel/Financial Sample.xlsx'
sheet='login_details'

a1 = get_cols_count(path,sheet)
print("Colums :", str(a1))

a2 = get_rows_count(path,sheet)
print("Rows ", str(a2))

a3 = get_data(path,sheet,2,2 )
print(a3)

a4 = set_data(path, sheet,4,2, "pass4")